"""
@author = allenchow
@email = 32502439@stu.hzcu.edu.cn
@current_time = 14/12/2025 12:08
"""
"""
@author = allenchow
@email = 32502439@stu.hzcu.edu.cn
@current_time = 13/12/2025 14:05
"""
import openpyxl
from openpyxl import load_workbook
import os
from datetime import datetime
import tkinter as tk
from tkinter import simpledialog, messagebox
import time
from tqdm import tqdm
# -------------------------- 核心配置（无需修改，已适配你的文件） --------------------------
EXCEL_FILE = "Stock_table.xlsx"  # 你的库存表路径（已定位到/mnt目录）
CHECKLIST_SHEET = "19-number"  # 核对表sheet名
STOCK_SHEET = "Stock"  # 库存主表shee9950015936522
# t名
LOG_SHEET = "log"  #日志sheet名
# 日志表头（固定）
LOG_HEADERS = ["操作时间", "条形码", "货号", "变动量", "操作类型", "操作前库存", "操作后库存", "操作状态", "备注"]
Stock_HEADERS = ["条码/19码","货号","商品名称","品牌","尺码","库存数量","备注"]
CHECKLIST_HEADERS = ["条码/19码","货号","商品名称","品牌","尺码","备注"]



def init_excel():
    """初始化Excel文件（如果不存在/缺少sheet/缺少表头，自动创建）"""

    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()

        # 核对表
        checklist = wb.active
        checklist.title = CHECKLIST_SHEET
        checklist.append(CHECKLIST_HEADERS)

        # 库存表
        stock = wb.create_sheet(STOCK_SHEET)
        stock.append(Stock_HEADERS)

        # 日志表
        log = wb.create_sheet(LOG_SHEET)
        log.append(LOG_HEADERS)

        wb.save(EXCEL_FILE)
        wb.close()
        return

    wb = openpyxl.load_workbook(EXCEL_FILE)

    # 检查核对表
    if CHECKLIST_SHEET not in wb.sheetnames:
        sheet = wb.create_sheet(CHECKLIST_SHEET)
        sheet.append(CHECKLIST_HEADERS)

    # 检查库存表
    if STOCK_SHEET not in wb.sheetnames:
        sheet = wb.create_sheet(STOCK_SHEET)
        sheet.append(Stock_HEADERS)

    # 检查日志表
    if LOG_SHEET not in wb.sheetnames:
        sheet = wb.create_sheet(LOG_SHEET)
        sheet.append(LOG_HEADERS)

    wb.save(EXCEL_FILE)
    wb.close()


def write_log(ws, barcode, sku, change, op_type, before, after, status, remark=""):
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        barcode,
        sku,
        change,
        op_type,
        before,
        after,
        status,
        remark
    ])

# def find_in_checklist(ws, barcode):
#     for row in range(3, ws.max_row + 1):
#         if str(ws.cell(row=row, column=1).value) == barcode:
#             return {
#                 "条码/19码": ws.cell(row, 1).value,
#                 "货号": ws.cell(row, 2).value,
#                 "商品名称": ws.cell(row, 3).value,
#                 "品牌": ws.cell(row, 4).value,
#                 "尺码": ws.cell(row, 5).value,
#                 "备注": ws.cell(row, 6).value
#             }
#     return None
#扫表模式效率低，改成内存索引

def build_index(ws, start_row):
    """
    构建 条码 -> 行号 的内存索引
    """
    index = {}
    for row in range(start_row, ws.max_row + 1):
        barcode = ws.cell(row, 1).value
        if barcode:
            index[str(barcode)] = row
    return index

def find_in_checklist(ws, checklist_index, barcode):
    row = checklist_index.get(barcode)
    if not row:
        return None

    return {
        "条码/19码": ws.cell(row, 1).value,
        "货号": ws.cell(row, 2).value,
        "商品名称": ws.cell(row, 3).value,
        "品牌": ws.cell(row, 4).value,
        "尺码": ws.cell(row, 5).value,
        "备注": ws.cell(row, 6).value
    }



#tk模式
# def manual_input(barcode):
#     root = tk.Tk()
#     root.withdraw()
#     data = {
#          "条码/19码": barcode,
#          "货号": simpledialog.askstring("录入货号", "请输入货号"),
#          "商品名称": simpledialog.askstring("录入商品名称", "请输入商品名称"),
#          "品牌": simpledialog.askstring("录入品牌", "请输入品牌"),
#          "尺码": simpledialog.askstring("录入尺码", "请输入尺码"),
#          "备注": "自建库" }
#     return data
def manual_input(barcode):
    print(f"条码 {barcode} 未在核对表中，开始建档，请依次输入信息：")

    sku = input("请输入货号: ").strip()
    if not sku:
        print("货号不能为空，建档取消")
        return None

    name = input("请输入商品名称: ").strip()
    if not name:
        print("商品名称不能为空，建档取消")
        return None

    brand = input("请输入品牌: ").strip()
    if not brand:
        print("品牌不能为空，建档取消")
        return None

    size = input("请输入尺码: ").strip()
    if not size:
        print("尺码不能为空，建档取消")
        return None

    data = {
        "条码/19码": barcode,
        "货号": sku,
        "商品名称": name,
        "品牌": brand,
        "尺码": size,
        "备注": "自建库"
    }

    print("建档完成！")
    return data






# def update_stock(ws, data):
#     for row in range(2, ws.max_row + 1):
#         if str(ws.cell(row, 1).value) == str(data["条码/19码"]):
#             before = ws.cell(row, 6).value or 0
#             ws.cell(row, 6, before + 1)
#             return before, before + 1
#
#     # 不存在，新增
#     ws.append([
#         data["条码/19码"],
#         data["货号"],
#         data["商品名称"],
#         data["品牌"],
#         data["尺码"],
#         1,
#         data["备注"]
#     ])
#     return 0, 1
#更新
def update_stock(ws, stock_index, data, mode):
    """
    mode: 'in' 入库, 'out' 出库
    """
    barcode = str(data["条码/19码"])

    if barcode not in stock_index:
        if mode == "out":
            return None, None, "not_exist"
        # 入库不存在时新增
        ws.append([
            data["条码/19码"],
            data["货号"],
            data["商品名称"],
            data["品牌"],
            data["尺码"],
            1,
            data["备注"]
        ])
        new_row = ws.max_row
        stock_index[barcode] = new_row
        return 0, 1, "success"

    row = stock_index[barcode]
    before = ws.cell(row, 6).value or 0

    if mode == "out":
        if before <= 0:
            return before, before, "no_stock"
        elif before == 1:
            # 出库后库存为0，删除行
            ws.delete_rows(row, 1)
            del stock_index[barcode]
            # 更新索引（删除行后，行号可能改变）
            for k, v in stock_index.items():
                if v > row:
                    stock_index[k] = v - 1
            return 1, 0, "deleted"
        else:
            ws.cell(row, 6, before - 1)
            return before, before - 1, "success"

    # 入库
    ws.cell(row, 6, before + 1)
    return before, before + 1, "success"



def scan_and_input(barcode):
    wb = load_workbook(EXCEL_FILE)
    checklist_ws = wb[CHECKLIST_SHEET]
    stock_ws = wb[STOCK_SHEET]
    log_ws = wb[LOG_SHEET]

    data = find_in_checklist(checklist_ws, barcode)

    if data:
        before, after = update_stock(stock_ws, data)
        write_log(log_ws, barcode, data["货号"], 1, "扫码入库", before, after, "成功")
    else:
        data = manual_input(barcode)

        # 写入 19-number
        checklist_ws.append([
            data["条码/19码"],
            data["货号"],
            data["商品名称"],
            data["品牌"],
            data["尺码"],
            data["备注"]
        ])

        before, after = update_stock(stock_ws, data)
        write_log(log_ws, barcode, data["货号"], 1, "自建入库", before, after, "成功", "自建库")

    wb.save(EXCEL_FILE)
    wb.close()
    print("录入成功！")


# def continuous_scan(mode):
#     """
#     mode: 'in' 入库, 'out' 出库
#     """
#     wb = load_workbook(EXCEL_FILE)
#     checklist_ws = wb[CHECKLIST_SHEET]
#     stock_ws = wb[STOCK_SHEET]
#     log_ws = wb[LOG_SHEET]
#
#     checklist_index = build_index(checklist_ws, 3)
#     stock_index = build_index(stock_ws, 2)
#
#     unknown_barcode = None
#
#     print(f"\n=== {'入库' if mode == 'in' else '出库'} 模式 ===")
#     print("扫码录入，输入 exit 退出并保存\n")
#
#     while True:
#         barcode = input("扫码：").strip()
#         if not barcode:
#             continue
#
#         if barcode.lower() == "exit":
#             break
#
#         data = find_in_checklist(checklist_ws, checklist_index, barcode)
#
#         # ---------- 条码不存在 ----------
#         if not data:
#             if mode == "out":
#                 print(f"⚠ 出库模式不允许未知条码：{barcode}")
#                 continue
#             unknown_barcode = barcode
#             print(f"⚠ 未知条码 {barcode}，退出扫码进入建档")
#             break
#
#         before, after, status = update_stock(
#             stock_ws, stock_index, data, mode
#         )
#
#         # ---------- 出库库存不足 ----------
#         if status == "no_stock":
#             print(f"⚠ 库存不足，无法出库：{barcode}")
#             continue
#         # ---------- 出库后删除库存 ----------
#         elif status == "deleted":
#             write_log(
#                 log_ws,
#                 barcode,
#                 data["货号"],
#                 -1,
#                 "出库",
#                 before,
#                 after,
#                 "成功",
#                 "库存归零已删除"
#             )
#             print(f"✔ {barcode} 出库完成，库存归零，已从库存表删除")
#             continue
#         # ---------- 普通入库/出库 ----------
#         write_log(
#             log_ws,
#             barcode,
#             data["货号"],
#             -1 if mode == "out" else 1,
#             "出库" if mode == "out" else "入库",
#             before,
#             after,
#             "成功"
#         )
#
#         print(f"✔ {barcode} {'出库' if mode == 'out' else '入库'}完成（{before} → {after}）")
#
#     wb.save(EXCEL_FILE)
#     wb.close()
#
#     return unknown_barcode if mode == "in" else None

# def continuous_scan(mode):
#     """
#     mode: 'in' 入库, 'out' 出库, 'check' 扫码检测模式
#     """
#     wb = load_workbook(EXCEL_FILE)
#     checklist_ws = wb[CHECKLIST_SHEET]
#     stock_ws = wb[STOCK_SHEET]
#     log_ws = wb[LOG_SHEET]
#
#     checklist_index = build_index(checklist_ws, 3)
#     stock_index = build_index(stock_ws, 2)
#
#     unknown_barcode = None
#
#     if mode == "in":
#         print("\n=== 入库模式 ===")
#         print("扫码录入，输入 exit 退出并保存\n")
#     elif mode == "out":
#         print("\n=== 出库模式 ===")
#         print("扫码录入，输入 exit 退出并保存\n")
#     elif mode == "check":
#         print("\n=== 扫码检测模式 ===")
#         print("扫码检测条码是否在核对表中，输入 exit 退出\n")
#         total_count = 0
#         exist_count = 0
#         not_exist_count = 0
#         not_exist_list = []
#
#     while True:
#         barcode = input("扫码：").strip()
#         if not barcode:
#             continue
#         if barcode.lower() == "exit":
#             break
#
#         if mode == "check":
#             total_count += 1
#             data = find_in_checklist(checklist_ws, checklist_index, barcode)
#             if data:
#                 exist_count += 1
#                 status = "存在"
#                 remark = f"第{total_count}个商品"
#                 print(f"✔ 第{total_count}个商品在核对表中")
#             else:
#                 not_exist_count += 1
#                 not_exist_list.append(total_count)
#                 status = "不存在"
#                 remark = f"第{total_count}个商品"
#                 print(f"⚠ 第{total_count}个商品未在核对表中")
#
#             # 写入日志
#             write_log(
#                 log_ws,
#                 barcode,
#                 data["货号"] if data else "-",
#                 0,  # 变动量为0
#                 "检测",
#                 "-" ,  # 操作前库存
#                 "-",  # 操作后库存
#                 status,
#                 remark
#             )
#             continue
#
#         # ---------- 条码不存在 ----------
#         data = find_in_checklist(checklist_ws, checklist_index, barcode)
#         if not data:
#             if mode == "out":
#                 print(f"⚠ 出库模式不允许未知条码：{barcode}")
#                 continue
#             unknown_barcode = barcode
#             print(f"⚠ 未知条码 {barcode}，退出扫码进入建档")
#             break
#
#         before, after, status = update_stock(stock_ws, stock_index, data, mode)
#
#         # ---------- 出库库存不足 ----------
#         if status == "no_stock":
#             print(f"⚠ 库存不足，无法出库：{barcode}")
#             continue
#         # ---------- 出库后删除库存 ----------
#         elif status == "deleted":
#             write_log(
#                 log_ws,
#                 barcode,
#                 data["货号"],
#                 -1,
#                 "出库",
#                 before,
#                 after,
#                 "成功",
#                 "库存归零已删除"
#             )
#             print(f"✔ {barcode} 出库完成，库存归零，已从库存表删除")
#             continue
#         # ---------- 普通入库/出库 ----------
#         write_log(
#             log_ws,
#             barcode,
#             data["货号"],
#             -1 if mode == "out" else 1,
#             "出库" if mode == "out" else "入库",
#             before,
#             after,
#             "成功"
#         )
#
#         print(f"✔ {barcode} {'出库' if mode == 'out' else '入库'}完成（{before} → {after}）")
#
#     wb.save(EXCEL_FILE)
#     wb.close()
#
#     if mode == "check":
#         print("\n=== 扫描统计结果 ===")
#         print(f"共扫描 {total_count} 个商品")
#         print(f"在核对表中的商品数：{exist_count}")
#         print(f"未在核对表中的商品数：{not_exist_count}")
#         if not_exist_list:
#             print(f"未在核对表中的商品是第 {not_exist_list} 个商品")
#
#     return unknown_barcode if mode == "in" else None


def continuous_scan(mode="in"):
    """
    mode:
    - in    : 入库
    - out   : 出库
    - check : 扫码检测模式（不改库存）
    """

    wb = load_workbook(EXCEL_FILE)
    checklist_ws = wb[CHECKLIST_SHEET]
    stock_ws = wb[STOCK_SHEET]
    log_ws = wb[LOG_SHEET]

    # 构建内存索引，提高查找效率
    checklist_index = build_index(checklist_ws, 3)
    stock_index = build_index(stock_ws, 2)

    total_count = 0
    exist_count = 0
    not_exist_map = {}  # 仅用于检测模式

    print(f"\n=== {'入库' if mode=='in' else '出库' if mode=='out' else '检测'} 模式 ===")
    print("扫码/输入条码，输入 exit 结束\n")

    unknown_barcode = None

    while True:
        barcode = input("扫码/输入条码：").strip()
        if not barcode:
            continue
        if barcode.lower() == "exit":
            break

        total_count += 1

        # 查找商品
        product = find_in_checklist(checklist_ws, checklist_index, barcode)

        # ---------- 检测模式 ----------
        if mode == "check":
            if product:
                exist_count += 1
                print(f"✔ 第{total_count}个商品在核对表中")
            else:
                print(f"⚠ 第{total_count}个商品未在核对表中")
                if barcode not in not_exist_map:
                    not_exist_map[barcode] = []
                not_exist_map[barcode].append(total_count)

            # 写日志（检测模式不改库存）
            write_log(
                log_ws,
                barcode,
                product["货号"] if product else "-",
                0,
                "检测",
                "-",
                "-",
                "存在" if product else "不存在",
                f"第{total_count}个商品"
            )
            continue

        # ---------- 入库 ----------
        if mode == "in":
            if product:
                before, after, status = update_stock(stock_ws, stock_index, product, "in")
                print(f"✔ {barcode} 入库完成（{before} → {after}）")
                write_log(
                    log_ws,
                    barcode,
                    product["货号"],
                    1,
                    "入库",
                    before,
                    after,
                    status,
                    ""
                )
            else:
                print(f"⚠ 未知条码 {barcode}，退出扫码进入建档")
                unknown_barcode = barcode
                break

        # ---------- 出库 ----------
        if mode == "out":
            if product:
                before, after, status = update_stock(stock_ws, stock_index, product, "out")
                if status == "no_stock":
                    print(f"⚠ 库存不足，无法出库：{barcode}")
                elif status == "deleted":
                    print(f"✔ {barcode} 出库完成，库存归零已删除")
                else:
                    print(f"✔ {barcode} 出库完成（{before} → {after}）")
                write_log(
                    log_ws,
                    barcode,
                    product["货号"],
                    -1,
                    "出库",
                    before,
                    after,
                    status,
                    ""
                )
            else:
                print(f"⚠ 条码不存在：{barcode}")
                write_log(
                    log_ws,
                    barcode,
                    "-",
                    -1,
                    "出库",
                    "-",
                    "-",
                    "失败",
                    "条码不存在"
                )

    # ---------- 检测模式结束 ----------
    if mode == "check":
        print("\n====== 扫码检测结果 ======")
        print(f"共扫描商品：{total_count}")
        print(f"存在于 19-number：{exist_count}")
        print(f"不存在于 19-number：{len(not_exist_map)}")
        write_log(
            log_ws,
            "",
            "",
            0,
            "扫码检测汇总",
            "-",
            "-",
            "完成",
            f"共{total_count}个，存在{exist_count}个，不存在{len(not_exist_map)}个"
        )
        wb.save(EXCEL_FILE)
        wb.close()
        return {
            "total": total_count,
            "exist": exist_count,
            "not_exist": len(not_exist_map),
            "not_exist_map": not_exist_map
        }

    wb.save(EXCEL_FILE)
    wb.close()
    return unknown_barcode  # 入库模式返回未知条码，用于建档



def batch_manual_input(barcodes):
    if not barcodes:
        print("没有需要建档的条码")
        return

    wb = load_workbook(EXCEL_FILE)
    checklist_ws = wb[CHECKLIST_SHEET]
    log_ws = wb[LOG_SHEET]

    for barcode in barcodes:
        choice = input(f"\n是否录入条码 {barcode} 的商品信息？(y/n): ").strip().lower()
        if choice != "y":
            print(f"跳过条码 {barcode}")
            continue

        data = manual_input(barcode)
        if not data:
            print("建档取消")
            continue

        checklist_ws.append([
            data["条码/19码"],
            data["货号"],
            data["商品名称"],
            data["品牌"],
            data["尺码"],
            data["备注"]
        ])

        write_log(
            log_ws,
            barcode,
            data["货号"],
            0,
            "建档",
            "-",
            "-",
            "成功",
            "模式4-补录"
        )

        print(f"✔ 条码 {barcode} 建档完成")

    wb.save(EXCEL_FILE)
    wb.close()



if __name__ == "__main__":
    init_excel()

    while True:
        print("\n请选择操作模式：")
        print("1 - 入库")
        print("2 - 出库")
        print("3 - 扫码检测模式（可补录 19-number）")
        print("0 - 退出程序")

        choice = input("请输入 0/1/2/3：").strip()

        # ---------- 模式 1：入库 ----------
        if choice == "1":
            unknown = continuous_scan(mode="in")
            if unknown:
                print(f"\n开始为条码 {unknown} 建档")
                data = manual_input(unknown)
                if data:
                    wb = load_workbook(EXCEL_FILE)
                    checklist_ws = wb[CHECKLIST_SHEET]
                    stock_ws = wb[STOCK_SHEET]
                    log_ws = wb[LOG_SHEET]

                    # 写入核对表
                    checklist_ws.append([
                        data["条码/19码"],
                        data["货号"],
                        data["商品名称"],
                        data["品牌"],
                        data["尺码"],
                        data["备注"]
                    ])

                    # 写入库存表
                    stock_ws.append([
                        data["条码/19码"],
                        data["货号"],
                        data["商品名称"],
                        data["品牌"],
                        data["尺码"],
                        1,
                        data["备注"]
                    ])

                    # 写入日志
                    write_log(
                        log_ws,
                        unknown,
                        data["货号"],
                        1,
                        "自建入库",
                        0,
                        1,
                        "成功",
                        "自建库"
                    )

                    wb.save(EXCEL_FILE)
                    wb.close()

        # ---------- 模式 2：出库 ----------
        elif choice == "2":
            continuous_scan(mode="out")

        # ---------- 模式 3：扫码检测 + 可进入补录 ----------
        elif choice == "3":
            result = continuous_scan(mode="check")

            # 如果没有未录入商品，直接回主菜单
            if not result or result["not_exist"] == 0:
                print("所有商品均在核对表中，无需补录")
                continue

            # 合并重复条码的索引，形成可选列表
            not_exist_items = []
            for barcode, indices in result["not_exist_map"].items():
                index_str = "、".join(str(idx) for idx in indices)
                not_exist_items.append({
                    "barcode": barcode,
                    "index": index_str
                })

            # 保存到 result 中方便后续使用
            result["not_exist_items"] = not_exist_items

            print("\n检测到存在未录入 19-number 的商品")
            go_fill = input("是否进入补录模式？(y/n)：").strip().lower()
            if go_fill != "y":
                continue

            # 循环补录，补录后从列表中删除已录条码
            while result["not_exist_items"]:
                print("\n未录入商品列表：")
                for i, item in enumerate(result["not_exist_items"], start=1):
                    print(f"{i}. 条码 {item['barcode']}（第 {item['index']} 个商品）")

                select = input("请输入要补录的编号（多个用逗号，如 1,3，输入0退出）：").strip()
                if select == "0":
                    break
                if not select:
                    continue

                selections = [int(x) for x in select.split(",") if x.isdigit()]
                wb = load_workbook(EXCEL_FILE)
                checklist_ws = wb[CHECKLIST_SHEET]
                log_ws = wb[LOG_SHEET]

                # 逆序删除已补录条码，避免索引错误
                selections = sorted(selections, reverse=True)
                for s in selections:
                    if 1 <= s <= len(result["not_exist_items"]):
                        barcode = result["not_exist_items"][s - 1]["barcode"]
                        print(f"\n开始补录条码 {barcode}")
                        data = manual_input(barcode)
                        if not data:
                            print("建档取消")
                            continue

                        checklist_ws.append([
                            data["条码/19码"],
                            data["货号"],
                            data["商品名称"],
                            data["品牌"],
                            data["尺码"],
                            data["备注"]
                        ])

                        write_log(
                            log_ws,
                            barcode,
                            data["货号"],
                            0,
                            "补录",
                            "-",
                            "-",
                            "成功",
                            "模式3检测后补录"
                        )

                        # 删除已补录条码
                        result["not_exist_items"].pop(s - 1)

                wb.save(EXCEL_FILE)
                wb.close()

        # ---------- 退出 ----------
        elif choice == "0":
            print("程序退出")
            break

        else:
            print("无效选择，请重新输入")

